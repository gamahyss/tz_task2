import os
import zipfile
import tempfile
import shutil
import rarfile
import py7zr
import docx
import openpyxl
import xlrd
import pdfplumber
import csv

TARGET_EXTS = {'doc', 'docx', 'xls', 'xlsx', 'pdf'} # список расширений рассматриваемых файлов
ARCHIVE_EXTS = {'zip', 'rar', '7z'} # список расширений рассматриваемых архивов

def extract_doc_docx(path):
    '''функция извлекает текст из документов с расширением .doc и .docx'''

    doc = docx.Document(path) # считываем содержимое документа
    text = []
    for para in doc.paragraphs: # добавляем каждый абзац в список элементов текста
        text.append(para.text)
    return '\n'.join(text) # преобразуем список в строку, разделённую на абзацы, и возвращаем это значение

def extract_xlsx(path):
    '''функция извлекает текст из документов с расширением .xlsx'''

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True) # загружаем документ, находящийся в указанной директории, в режиме чтения и без формул
    text = []
    for sheet in wb.worksheets: # перебираем листы, в каждом листе перебираем строки (рассматриваем только значения без информации о стилях и т.д.)
        for row in sheet.iter_rows(values_only=True):
            for cell in row: # в каждой строке перебираем клетки и добавляем данные содержащиеся в клетке в список элементов текста (если эти данные вообще есть)
                if cell is not None:
                    text.append(str(cell))
    return '\n'.join(text) # преобразуем список элементов в строку из этих элементов, разделённых абзацами, и возвращаем

def extract_xls(path):
    '''функция извлекает текст из документов с расширением .xls'''

    wb = xlrd.open_workbook(path) # считываем содержимое документа
    text = []
    for sheet in wb.sheets(): # перебираем листы экселя, в каждом листе перебираем строки
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols): # в каждой строке перебираем колонки
                cell = sheet.cell(row_idx, col_idx) # получаем клетку по индексу строки и индексу колонки
                if cell.ctype != xlrd.XL_CELL_EMPTY: # рассматриваем текущую клетку только если она не пустая
                    if cell.ctype == xlrd.XL_CELL_DATE: # если тип данных в клетке - дата, то преобразуем данные в кортеж, а его - в строку и добавляем в список содержимого
                        date_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                        text.append(str(date_tuple))
                    else: # если тип данных другой, преобразуем данные в строку и добаляем в список содержимого
                        text.append(str(cell.value))
    return '\n'.join(text) # преобразуем список содержимого в строку, где элементы разделены абзацами, и возвращаем

def extract_pdf(path):
    '''функция извлекает текст из документов с расширением .pdf'''

    text = []
    with pdfplumber.open(path) as pdf: # открываем с помощью библиотеки pdfplumber
        for page in pdf.pages: # перебираем все страницы и извлекаем текст с каждой
            page_text = page.extract_text()
            if page_text: # если текст не пустой то добавляем в список содержимого
                text.append(page_text)
    return '\n'.join(text) # преобразуем список содержимого в строку, где элементы разделены абзацами, и возвращаем

EXTRACTORS = { # соответствие расширений файлов тем функциям, которые с ними работают
    'docx': extract_doc_docx,
    'doc': extract_doc_docx,
    'xlsx': extract_xlsx,
    'xls': extract_xls,
    'pdf': extract_pdf,
}

def process_file(full_path, storage_path):
    '''функция для обработки какого-либо файла'''

    ext = os.path.splitext(full_path)[1].lower().lstrip('.') # находим расширение файла
    if ext in EXTRACTORS: # если оно находится в нужных нам расширениях, то используем соответствующую функцию для извлечения текста
        text = EXTRACTORS[ext](full_path)
        return { # возвращаем данные в формате: путь к файлу, имя файла, расширение, содержимое
            'file_path': storage_path,
            'file_name': os.path.basename(full_path),
            'extension': ext,
            'content': text
        }
    return None # если у файла какое-то другое расширение, возвращаем None

def process_archive(archive_path, archive_storage_path):
    '''функция для обработки архива'''

    ext = os.path.splitext(archive_path)[1].lower().lstrip('.') # получаем расширение архива
    temp_dir = tempfile.mkdtemp() # создаём времменную директорию
    results = []
    try:
        if ext == 'zip': # распаковываем архив с помощью той библиотеки которая работает с его расширением
            with zipfile.ZipFile(archive_path, 'r') as zf:
                zf.extractall(temp_dir)
        elif ext == 'rar':
            with rarfile.RarFile(archive_path) as rf:
                rf.extractall(temp_dir)
        elif ext == '7z':
            with py7zr.SevenZipFile(archive_path, mode='r') as sz:
                sz.extractall(temp_dir)
        for root, dirs, files in os.walk(temp_dir): # перебираем путь, директории, файлы во временной директории
            rel_dir = os.path.relpath(root, temp_dir) # находим относительный путь к указанному пути
            if rel_dir == '.': # если это текущий каталог то заменяем на пустую строку
                rel_dir = ''
            for file in files: # перебираем каждый файл
                file_full = os.path.join(root, file) # находим полный путь к файлу
                inner_rel = os.path.join(rel_dir, file) if rel_dir else file # находим внутренний относительный путь
                storage_subpath = os.path.join(archive_storage_path, inner_rel).replace('\\', '/') # соединяем путь к архиву и относительный путь внутри архива
                file_ext = os.path.splitext(file)[1].lower().lstrip('.') # находим расширение файла
                if file_ext in ARCHIVE_EXTS: # если это архив в архиве, то внутренний архив обрабатываем с помощью этой же функции
                    sub_results = process_archive(file_full, storage_subpath)
                    results.extend(sub_results)
                elif file_ext in TARGET_EXTS: # если это файл с расширением из списка нужных нам расширений, то обрабатываем файл и добавляем в результат
                    res = process_file(file_full, storage_subpath)
                    if res:
                        results.append(res)
    except Exception as e: # ловим exception и выводим информацию о нём при ошибке обработки
        print(f"Ошибка при обработке архива {archive_path}: {e}")
    finally: # после всех действий убираем временную директорию, игнорируя ошибки
        shutil.rmtree(temp_dir, ignore_errors=True)
    return results


def crawl_storage(root_path):
    '''функция для краулинга всего хранилища'''

    results = []
    root_path = os.path.abspath(root_path) # находим абсолютный путь к хранилищу

    for dirpath, dirnames, filenames in os.walk(root_path): # перебираем путь, директории, файлы в хранилище
        for filename in filenames: # перебираем каждый файл
            full_path = os.path.join(dirpath, filename) # находим абсолютный путь к файлу
            rel_path = os.path.relpath(full_path, root_path).replace('\\', '/') # находим относительный путь
            ext = os.path.splitext(filename)[1].lower().lstrip('.') # находим расширение файла

            if ext in TARGET_EXTS: # если у файла то расширение которое нам нужно, то обрабатываем файл
                res = process_file(full_path, rel_path)
                if res: # если содержимое не пустое, то добавляем в результат
                    results.append(res)
            elif ext in ARCHIVE_EXTS: # если у файла расширение одного из нужных нам архивов, то обрабатываем как архив и тоже добавляем в результат
                arc_results = process_archive(full_path, rel_path)
                results.extend(arc_results)
    return results

if __name__ == '__main__':
    '''главный блок программы'''

    while True: # сначала спрашиваем у пользователя путь к хранилищу
        root = input("Введите путь к папке для сканирования: ").strip()
        if not root:
            print("Путь не может быть пустым.")
            continue
        if not os.path.isdir(root):
            print("Указанный путь не существует или не является папкой.")
        break

    print("Сканирование...")
    results = crawl_storage(root) # сканируем файлы из полученной папки

    if not results:
        print("Не найдено ни одного поддерживаемого файла.")

    print(f"Найдено файлов: {len(results)}")

    out_file = input(f"Имя выходного CSV-файла БЕЗ РАСШИРЕНИЯ (по умолчанию 'output.csv'): ").strip() # опрашиваем имя файла, в который надо вывести результат
    if not out_file:
        out_file = 'output.csv'
    else:
        out_file = out_file + '.csv'

    with open(out_file, 'w', encoding='utf-8', newline='') as f: # открываем csv файл с кодировкой utf-8 в режиме записи
        writer = csv.writer(f, quoting=csv.QUOTE_ALL) # объект который будет записывать результаты в файл
        writer.writerow(['file_path', 'file_name', 'extension', 'content']) # сначала пишем названия колонок с информацией
        for item in results: # затем заполняем каждую колонку
            writer.writerow([
                item['file_path'],
                item['file_name'],
                item['extension'],
                item['content']
            ])
    print(f"Результаты сохранены в '{out_file}'")